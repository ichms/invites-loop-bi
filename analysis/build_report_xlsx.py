"""data/csv/ 의 CSV 를 미션(01) 기준으로 합쳐 공유용 xlsx 한 파일로 만듭니다.

    uv run python analysis/build_report_xlsx.py

노트북(point_mission_liability_ko.ipynb)과 같이 전달하는 것을 전제로 하며,
시트 구성과 `판단` 컬럼은 노트북 4·6절과 같은 계산을 씁니다.

**사용자 단위 CSV(02/03/04/08)는 넣지 않습니다.** raw user_no / user_id(UUID)가 들어 있어
외부 공유 대상이 아니기 때문입니다. 미션 단위로 집계한 값만 담습니다.
"""

from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REF_DATE = "2026-08-04"
WINDOW = "2026-07-06 ~ 2026-08-04"


def find_data() -> Path:
    for base in [Path.cwd(), *Path.cwd().parents]:
        if (base / "data" / "csv" / "01_mission_catalog.csv").exists():
            return base / "data" / "csv"
    raise FileNotFoundError("data/csv 를 찾지 못했습니다. `bash data/sql/build_all.sh` 를 먼저 실행하세요.")


DATA = find_data()
OUT = DATA.parent / "포인트_미션_분석.xlsx"

catalog   = pd.read_csv(DATA / "01_mission_catalog.csv")
users     = pd.read_csv(DATA / "02_user_activity_30d.csv")
earn      = pd.read_csv(DATA / "03_mission_user_earn_30d.csv")
summary   = pd.read_csv(DATA / "05_mission_activity_summary.csv")
liability = pd.read_csv(DATA / "06_point_liability_by_mission.csv")
monthly   = pd.read_csv(DATA / "07_mission_monthly_issue.csv", parse_dates=["month"])

LABEL = {
    1: "회원가입", 2: "성향문진", 3: "글쓰기(30자+)", 5: "좋아요 10개",
    6: "연속 글쓰기 3일", 7: "연속 글쓰기 5일", 8: "연속 글쓰기 7일", 9: "AI 글쓰기",
    43: "댓글 쓰기", 10: "애드센스 연동", 11: "게임 실행", 12: "게임 랭킹 1위",
    13: "게임 랭킹 2~10위", 14: "출석 1·2·4·5일차", 15: "출석 3·6일차", 16: "출석 7일차",
    17: "글 공유 유입", 18: "유형카드 공유 유입", 19: "첫 친구 초대", 20: "친구 초대 2~19명",
    21: "친구 초대 5명", 22: "친구 초대 10·20명", 23: "친구 초대 15명",
    24: "친구 초대 21~50명", 25: "친구 초대 51~100명", 26: "친구 초대 101~1000명",
    34: "생활문진 업데이트", 35: "첫 7일 연속 루틴(삭제됨)", 36: "7일 연속 루틴 수행",
    37: "루틴 1개 응답", 38: "식단 등록", 39: "일일 피드백", 40: "가족 등록 수락",
    41: "12주 완주", 42: "만족도 조사", 101: "관리자 수동 지급",
    102: "지타이 개인화 시작", 103: "지타이 기준선 생성", 104: "기분 기록 1~6일",
    105: "기분 기록 7일", 106: "기분 기록 8~13일", 107: "기분 기록 14일",
    108: "기분 기록 15일+", 109: "주간 체크업",
}


def label(no) -> str:
    return "(미분류)" if pd.isna(no) else LABEL.get(int(no), f"미션 #{int(no)}")


# ── 노트북 4·6절과 동일한 효과 추정 ─────────────────────────────────────────
DISCRETIONARY = {101}          # 미션이 아니라 관리자 재량 지급
OUTCOME_OVERLAP = {35, 36, 37, 39}   # 지급 조건 자체가 루틴 응답 = 성과 지표


def estimate_effects() -> pd.DataFrame:
    pop = users[users.routine_assigned_30d > 0].dropna(
        subset=["routine_done_rate_30d", "visit_days_30d"])
    rows = []
    for no, uset in earn.groupby("point_mission_no").user_no.apply(set).items():
        d = pop.assign(joined=pop.user_no.isin(uset).astype(int))
        if d.joined.sum() < 20 or (1 - d.joined).sum() < 20:
            continue
        m = sm.OLS(d.routine_done_rate_30d,
                   sm.add_constant(d[["joined", "visit_days_30d"]])).fit()
        a = d.loc[d.joined == 1, "routine_done_rate_30d"]
        b = d.loc[d.joined == 0, "routine_done_rate_30d"]
        rows.append({"point_mission_no": no, "적립자수(효과추정)": int(d.joined.sum()),
                     "단순 차이": a.mean() - b.mean(),
                     "조정 계수": m.params["joined"], "p(조정)": m.pvalues["joined"]})
    return pd.DataFrame(rows)


def verdict(r) -> str:
    if r["30일 적립자"] == 0:
        return "30일 미발생"
    if pd.isna(r["조정 계수"]):
        return "표본 부족 — 판단 불가"
    if r["미션번호"] in DISCRETIONARY:
        return "통제 대상 — 재량 지급(인과 역전 의심)"
    if r["미션번호"] in OUTCOME_OVERLAP:
        return "판단보류 — 지급 조건이 성과 지표와 겹침"
    if r["조정 계수"] > 0 and r["p(조정)"] < 0.05:
        return "유지"
    if r["30일 발행"] >= 50_000:
        return "축소 검토 — 고비용·효과 불명"
    return "관망 — 저비용"


# ── 시트 1: 미션 종합 (01 기준으로 전부 조인) ───────────────────────────────
solo = set((s := earn.groupby("user_no").point_mission_no.nunique())[s == 1].index)
overlap = (earn.groupby("point_mission_no")
           .agg(전속_적립자=("user_no", lambda x: len(set(x) & solo)))
           .reset_index())

master = (catalog.assign(미션=catalog.point_mission_no.map(label))
          .merge(liability[["point_mission_no", "points_issued", "points_outstanding",
                            "points_used", "points_issued_30d", "points_issued_per_user"]],
                 on="point_mission_no", how="left")
          .merge(summary[["point_mission_no", "avg_age", "female_ratio",
                          "avg_visit_days_30d", "avg_routine_done_rate_30d",
                          "avg_routine_fdbk_cnt_30d", "avg_survey_cnt_30d",
                          "pct_used_point_mall_30d"]],
                 on="point_mission_no", how="left")
          .merge(overlap, on="point_mission_no", how="left"))

master = master.rename(columns={
    "point_mission_no": "미션번호", "description": "미션 설명", "mission_type": "유형",
    "parent_category_name": "상위 카테고리", "category_name": "카테고리",
    "use_yn": "사용", "amount": "지급액", "limit_type": "제한유형", "limit_cnt": "제한횟수",
    "expire_day_cnt": "소멸일수",
    "points_issued": "누적 발행", "points_outstanding": "미상환 잔액", "points_used": "누적 사용",
    "points_issued_30d": "30일 발행", "points_issued_per_user": "1인당 발행",
    "total_unique_users": "누적 적립자", "earners_30d": "30일 적립자",
    "earn_rows_30d": "30일 적립 건수", "earners_30d_loop": "30일 적립자(LOOP)",
    "avg_age": "적립자 평균나이", "female_ratio": "적립자 여성비율",
    "avg_visit_days_30d": "적립자 평균 방문일", "avg_routine_done_rate_30d": "적립자 루틴 수행률",
    "avg_routine_fdbk_cnt_30d": "적립자 루틴 피드백", "avg_survey_cnt_30d": "적립자 생활문진",
    "pct_used_point_mall_30d": "적립자 포인트몰 사용률",
})
master = master.merge(estimate_effects(), left_on="미션번호",
                      right_on="point_mission_no", how="left").drop(columns="point_mission_no")
master["효과 1%p당 비용"] = np.where(master["조정 계수"] > 0,
                                 master["30일 발행"] / (master["조정 계수"] * 100), np.nan)
master["판단"] = master.apply(verdict, axis=1)
master["전속 적립자"] = master.전속_적립자.fillna(0).astype(int)

MASTER_COLS = [
    "미션번호", "미션", "미션 설명", "유형", "상위 카테고리", "카테고리",
    "사용", "지급액", "제한유형", "제한횟수", "소멸일수",
    "누적 발행", "미상환 잔액", "누적 사용", "누적 적립자", "1인당 발행",
    "30일 발행", "30일 적립자", "30일 적립 건수", "30일 적립자(LOOP)", "전속 적립자",
    "적립자 평균나이", "적립자 여성비율", "적립자 평균 방문일", "적립자 루틴 수행률",
    "적립자 루틴 피드백", "적립자 생활문진", "적립자 포인트몰 사용률",
    "단순 차이", "조정 계수", "p(조정)", "효과 1%p당 비용", "판단",
]
master = master[MASTER_COLS].sort_values("30일 발행", ascending=False)

# ── 나머지 시트 ────────────────────────────────────────────────────────────
liab_sheet = (liability.assign(미션=liability.point_mission_no.map(label))
              .rename(columns={"point_mission_no": "미션번호", "mission_description": "미션 설명",
                               "mission_use_yn": "사용", "mission_del_yn": "삭제",
                               "mission_amount": "지급액", "users_all_time": "누적 적립자",
                               "points_issued": "누적 발행", "points_outstanding": "미상환 잔액",
                               "points_used": "누적 사용", "points_returned": "반환",
                               "points_issued_per_user": "1인당 발행",
                               "redemption_rate": "상환율", "lots_30d": "30일 건수",
                               "points_issued_30d": "30일 발행",
                               "points_issued_30d_loop": "30일 발행(LOOP)",
                               "last_issue_at": "마지막 발행"})
              [["미션번호", "미션", "미션 설명", "사용", "삭제", "지급액", "누적 적립자",
                "누적 발행", "미상환 잔액", "누적 사용", "반환", "1인당 발행",
                "30일 건수", "30일 발행", "30일 발행(LOOP)", "마지막 발행"]]
              .sort_values("누적 발행", ascending=False))

mon = (monthly.assign(미션=monthly.point_mission_no.map(label),
                      월=monthly.month.dt.strftime("%Y-%m"))
       .pivot_table(index="미션", columns="월", values="points_issued", aggfunc="sum")
       .fillna(0).astype(int))
mon["합계"] = mon.sum(axis=1)
mon = mon.sort_values("합계", ascending=False).reset_index()

act = (summary.assign(미션=summary.point_mission_no.map(label))
       .rename(columns={"point_mission_no": "미션번호", "mission_description": "미션 설명",
                        "n_earners_30d": "30일 적립자", "avg_age": "평균나이",
                        "female_ratio": "여성비율", "avg_visit_days_30d": "평균 방문일",
                        "avg_login_days_30d": "평균 접속일", "avg_routine_done_30d": "평균 루틴 수행",
                        "avg_routine_done_rate_30d": "루틴 수행률",
                        "avg_routine_no_response_30d": "루틴 무응답",
                        "avg_routine_fdbk_cnt_30d": "루틴 피드백",
                        "avg_survey_cnt_30d": "생활문진 횟수",
                        "med_survey_median_gap_days": "생활문진 주기(일)",
                        "pct_used_point_mall_30d": "포인트몰 사용률",
                        "pct_posted_30d": "게시글 작성률", "pct_commented_30d": "댓글 작성률",
                        "avg_points_earned_30d": "1인당 총 적립P"})
       [["미션번호", "미션", "미션 설명", "30일 적립자", "평균나이", "여성비율",
         "평균 방문일", "평균 접속일", "평균 루틴 수행", "루틴 수행률", "루틴 무응답",
         "루틴 피드백", "생활문진 횟수", "생활문진 주기(일)", "포인트몰 사용률",
         "게시글 작성률", "댓글 작성률", "1인당 총 적립P"]]
       .sort_values("30일 적립자", ascending=False))

issued, outstanding = liability.points_issued.sum(), liability.points_outstanding.sum()
used, returned = liability.points_used.sum(), liability.points_returned.sum()
issued_30d, used_30d = liability.points_issued_30d.sum(), users.point_used_30d.sum()

overview = pd.DataFrame({
    "항목": ["누적 발행", "누적 사용(상환)", "반환", "미상환 잔액 (= 부채)",
             "최근 30일 발행", "최근 30일 사용", "최근 30일 순증"],
    "포인트": [issued, used, returned, outstanding, issued_30d, used_30d, issued_30d - used_30d],
})
overview["총 발행 대비"] = overview.포인트 / issued

COVER = [
    ("포인트 미션 분석 — 요약 데이터", ""),
    ("", ""),
    ("기준일", REF_DATE),
    ("30일 윈도우", WINDOW),
    ("출처", "invites_dw 웨어하우스 · data/sql/build_all.sh 로 재생성"),
    ("함께 보는 문서", "analysis/point_mission_liability_ko.ipynb (해석과 그래프)"),
    ("", ""),
    ("시트 안내", ""),
    ("  1. 부채 요약", "포인트 발행·사용·잔액 전체 집계"),
    ("  2. 미션 종합", "미션 1행. 카탈로그·부채·적립자 활동·효과 추정을 모두 합친 메인 시트"),
    ("  3. 포인트 부채", "미션별 발행/잔액/사용 금액 상세"),
    ("  4. 월별 발행", "미션 × 월 발행액 추이"),
    ("  5. 적립자 활동", "미션별 적립자 집단의 30일 활동 프로필"),
    ("", ""),
    ("읽을 때 주의", ""),
    ("  조정 계수", "접속 일수가 비슷한 사람끼리 비교했을 때 남는 루틴 수행률 차이"),
    ("  p(조정)", "0.05 보다 작으면 우연으로 보기 어려운 차이"),
    ("  판단", "인과가 아니라 연관입니다. 미션을 실제로 껐다 켜야 확인됩니다"),
    ("  판단보류", "지급 조건이 곧 성과 지표인 미션은 이 방식으로 평가할 수 없습니다"),
    ("", ""),
    ("", ""),
    ("합계가 시트마다 다른 이유", ""),
    ("  시트 2 vs 시트 3", "시트 2는 카탈로그(삭제되지 않은 미션 43개) 기준이라 이미 삭제된 미션 #35"),
    ("", "와 미션이 지정되지 않은 지급분이 빠집니다. 전체 금액은 시트 3을 보십시오."),
    ("  30일 발행 vs 30일 적립자", "'발행'은 원장 전체 금액, '적립자'는 적립 판정을 통과한 건만 셉니다."),
    ("", "그래서 적립자 수 × 지급액이 발행액과 딱 맞지 않습니다."),
    ("", ""),
    ("포함하지 않은 것", "사용자 단위 데이터(02/03/04/08)는 개인 식별자가 있어 제외했습니다"),
]

# ── 쓰기 + 서식 ────────────────────────────────────────────────────────────
PCT = {"적립자 여성비율", "적립자 루틴 수행률", "적립자 포인트몰 사용률", "여성비율",
       "루틴 수행률", "포인트몰 사용률", "게시글 작성률", "댓글 작성률", "총 발행 대비", "상환율"}
INT = {"누적 발행", "미상환 잔액", "누적 사용", "반환", "30일 발행", "30일 발행(LOOP)",
       "1인당 발행", "포인트", "1인당 총 적립P", "효과 1%p당 비용", "합계"}
DEC3 = {"단순 차이", "조정 계수"}

HEAD_FILL = PatternFill("solid", fgColor="2F3437")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")

with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
    pd.DataFrame(COVER, columns=["항목", "내용"]).to_excel(xl, sheet_name="0. 읽는 법", index=False, header=False)
    overview.to_excel(xl, sheet_name="1. 부채 요약", index=False)
    master.to_excel(xl, sheet_name="2. 미션 종합", index=False)
    liab_sheet.to_excel(xl, sheet_name="3. 포인트 부채", index=False)
    mon.to_excel(xl, sheet_name="4. 월별 발행", index=False)
    act.to_excel(xl, sheet_name="5. 적립자 활동", index=False)

    for ws in xl.book.worksheets:
        name = ws.title
        if name.startswith("0."):
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 78
            ws["A1"].font = Font(bold=True, size=13)
            continue
        headers = [c.value for c in ws[1]]
        for c in ws[1]:
            c.fill, c.font = HEAD_FILL, HEAD_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "C2" if "미션" in headers else "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 30
        for i, h in enumerate(headers, start=1):
            col = get_column_letter(i)
            width = 34 if h in ("미션 설명", "판단") else (16 if h == "미션" else
                    max(9, min(15, len(str(h)) * 2)))
            ws.column_dimensions[col].width = width
            fmt = ("0.0%" if h in PCT else "#,##0" if h in INT else
                   "0.000" if h in DEC3 else "0.00E+00" if h == "p(조정)" else None)
            if fmt:
                for cell in ws[col][1:]:
                    cell.number_format = fmt
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.border = Border(bottom=THIN)

print(f"생성 완료: {OUT}")
print(f"  시트 6개 · 미션 종합 {len(master)}행 × {len(MASTER_COLS)}열")
